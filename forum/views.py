from django.contrib.auth.models import User
from django.http import HttpResponse

from rest_framework.permissions import IsAdminUser
from rest_framework.views import APIView
from rest_framework.decorators import action
# from django.core.exceptions import ObjectDoesNotExist
from openpyxl.styles import Font  
import openpyxl
from rest_framework import viewsets 
from rest_framework.exceptions import PermissionDenied,NotFound
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination

from rest_framework.viewsets import GenericViewSet
from rest_framework.mixins import CreateModelMixin,RetrieveModelMixin,UpdateModelMixin,DestroyModelMixin
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.shortcuts import get_object_or_404
from .models import (
    Question, Answer, QuestionVote,
    AnswerVote,UserProfile,
    Report
    )
# from .permissions import IsAdminOrReadOnly,IsOwnerOrAdmin
from .serializers import (
    QuestionSerializer, AnswerSerializer, 
    QuestionVoteSerializer, AnswerVoteSerializer,
    ReportSerializer,
    UserSerializer,UserProfileSerializer
)
from .permissions import (
    IsOwnerOrAdmin,IsOwnerOrAdminForAnswer,IsAuthenticatedAndCanVote,CanGenerateOrViewReport
)



class UserProfileViewset(viewsets.ModelViewSet):
    permission_classes=[IsAdminUser]
    queryset=UserProfile.objects.all()
    pagination_class=PageNumberPagination
    serializer_class=UserProfileSerializer

    # # breakpoint()
    # serializer_class=UserProfileSerializer 
    
    
    def get_queryset(self):
        user=self.request.user
        
        if user.is_authenticated and user.profile.role == "ADMIN":
            return UserProfile.objects.all()
        # breakpoint()
        # print("kjdfhaksdfdsjfjas")
        return UserProfile.objects.get(id=user.id)
        
        
    
    # def perform_create(self, serializer):
    #     serializer.save()
    

    


class QuestionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class=PageNumberPagination
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    
    
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser], url_path='download-report')
    def download_report(self, request, pk=None):
        
        try:
            question = Question.objects.get(id=pk)
        except Question.DoesNotExist:
            return Response({"error": "Question not found."}, status=404)

        
        answers = Answer.objects.filter(question=question)

        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Question {pk}"

        
        question_cell = ws.cell(row=1, column=1, value=f"Question: {question.title}")
        question_cell.font = Font(bold=True)

        
        ws.append([])

        
        ws.append(["Answer ID", "Answer Text", "Answered By", "Created At"])

        
        for answer in answers:
            ws.append([
                answer.id,
                answer.body,
                answer.answered_by.username,
                answer.answered_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="question_{question.id}_report.xlsx"'

        
        wb.save(response)

        return response

    def get_queryset(self):
        user = self.request.user

        if not user.is_authenticated:  
            raise PermissionDenied({"message": "Please Login First!"})

        if user.profile.role == "ADMIN":
            return Question.objects.all()
        elif user.profile.role == "USER":
            return Question.objects.filter(Q(status="APPROVED") | Q(created_by=user)).order_by('-created_at')

        return Question.objects.none()
    
    def destroy(self, request, *args, **kwargs):
        question = self.get_object()
        user=request.user

        if question.status == "APPROVED"  and not user.is_staff:
            raise PermissionDenied({"message": "Approved questions cannot be deleted!"})
        if question.status == "PENDING" and question.created_by != user:
            raise PermissionDenied({"message": "You can only delete your own pending questions!"})
        return super().destroy(request, *args, **kwargs)

class AdminDownloadQuestionReportView(APIView):
    permission_classes = [IsAdminUser]  

    def get(self, request, question_id):

        try:
            question = Question.objects.get(id=question_id)
        except Question.DoesNotExist:
            return Response({"error": "Question not found."}, status=404)

        
        answers = Answer.objects.filter(question=question)

        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Question {question_id}"

        
        question_cell = ws.cell(row=1, column=1, value=f"Question: {question.text}")
        question_cell.font = Font(bold=True)

        
        ws.append([])

        
        ws.append(["Answer ID", "Answer Text", "Answered By", "Created At"])

        
        for answer in answers:
            ws.append([
                answer.id,
                answer.text,
                answer.user.username,
                answer.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="question_{question.id}_report.xlsx"'

        
        wb.save(response)

        return response
            



class AnswerViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = AnswerSerializer

    def get_queryset(self):
        question_id = self.kwargs["question_pk"]
        question = get_object_or_404(Question, id=question_id)

        if question.status != "APPROVED":
            raise NotFound("You are not authorized to view answers for this question.")

        return Answer.objects.filter(question=question)

    def perform_create(self, serializer):
        question_id = self.kwargs.get("question_pk")
        question = get_object_or_404(Question, id=question_id)
        serializer.save(question=question, answered_by=self.request.user) 
        
    def destroy(self, request, *args, **kwargs):
        answer = self.get_object()
        user=request.user
        if not answer.answered_by == user:
            raise PermissionDenied({"message": "You can't delete this answer"})
        return super().destroy(request, *args, **kwargs) 
        

class QuestionVoteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticatedAndCanVote]
    queryset=QuestionVote.objects.all()
    serializer_class = QuestionVoteSerializer
    
    def get_queryset(self):
        user=self.request.user

        
        question = get_object_or_404(Question, id=self.kwargs["question_pk"])
        
        votes= question.q_votes.all()
        
        if not question.status=="APPROVED":
            raise PermissionDenied({"message":"question is not approved yet to vote"})
        
    
        # elif question.status=="APPROVED":
            
            
            
        # elif not votes.exists():
        #     raise PermissionDenied({"message": "No votes till date"})
        

        return votes
        # queryset=QuestionVote.objects.filter(question__id=int(question_id))
        
        
        # return queryset
    

    # def get_queryset(self):
    #     question_id = self.kwargs["pk"]  
    #     return QuestionVote.objects.filter(question_id=question_id)  


class AnswerVoteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticatedAndCanVote]
    queryset=AnswerVote.objects.all()
    serializer_class = AnswerVoteSerializer
    # def get_queryset(self):
    #     question_pk = self.kwargs["question_pk"]  
    #     answer_pk = self.kwargs["answer_pk"]  
    #     return AnswerVote.objects.filter(answer_id=answer_pk, answer__question_id=question_pk)  
    

class ReportViewset(viewsets.ModelViewSet):
    permission_classes=[IsAuthenticated,CanGenerateOrViewReport,IsOwnerOrAdmin]
    queryset=Report.objects.all()
    
    
    serializer_class=ReportSerializer
    
    # def get_renderers(self):
        
    #     if self.action == "list":
    #         return [JSONRenderer()]
    #     return super().get_renderers()
    def get_queryset(self):
        requester=(self.request.user)
        # breakpoint()
        # req_id=requester.id
        # breakpoint()
        # breakpoint()
        # print(type(user))
        if requester.profile.role=="ADMIN":
            # return Report.objects.filter(status="PENDING").order_by("-created_at")
            return Report.objects.all()
        
        return Report.objects.filter(user=requester).order_by("-created_at")


class UserViewset(viewsets.ModelViewSet):
    
    queryset=User.objects.all()
    serializer_class=UserSerializer
    
    def get_queryset(self):
        user=self.request.user
        if user.is_staff:
            return User.objects.all()
        elif not user.is_staff:
            return User.objects.filter(id=user.id)
            
            
            
        
    